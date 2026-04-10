"""
Build and maintain a database of publications.
"""

from __future__ import print_function, division, unicode_literals

# Standard library
import os
import re
import sys
import json
import datetime
import argparse
import collections
import numpy as np
import yaml
import requests
import readline
import webbrowser
from pprint import pprint
import logging
import jinja2
import pandas as pd
import pdb
from db_mongo_conn import MongoDBConnector 
#init logging
logging.basicConfig(level=logging.INFO)
log = logging.getLogger('KPUB')
log.setLevel(logging.INFO)
try:    
    import textract
except: 
    textract = None
    log.error("Could not import textract!  Will not be able to parse PDF text.")

import plot


#misc globals
PACKAGEDIR = os.path.abspath(os.path.dirname(__file__))
PLOTDIR = f"{PACKAGEDIR}/../data/plots"
MDDIR   = f"{PACKAGEDIR}/../data/output"


#ADS API URL
ADS_API = 'https://api.adsabs.harvard.edu/v1/search/query?'

# Which metadata fields do we want to retrieve from the ADS API?
# (basically everything apart from 'body' to reduce data volume)
FIELDS = ['date', 'pub', 'id', 'volume', 'links_data', 'citation', 'doi',
          'eid', 'keyword_schema', 'citation_count', 'data', 'data_facet',
          'year', 'identifier', 'keyword_norm', 'reference', 'abstract', 'recid',
          'alternate_bibcode', 'arxiv_class', 'bibcode', 'first_author_norm',
          'pubdate', 'reader', 'doctype', 'doctype_facet_hier', 'title', 'pub_raw', 'property',
          'author', 'email', 'orcid', 'keyword', 'author_norm',
          'cite_read_boost', 'database', 'classic_factor', 'ack', 'page',
          'first_author', 'reader', 'read_count', 'indexstamp', 'issue', 'keyword_facet',
          'aff', 'facility', 'simbid']

#Defines colors for highlighting words in the terminal.
HIGHLIGHTS = {
    "RED"    : "\033[4;31m",
    "GREEN"  : "\033[4;32m",
    "YELLOW" : "\033[4;33m",
    "BLUE"   : "\033[4;34m",
    "PURPLE" : "\033[4;35m",
    "CYAN"   : "\033[4;36m",
    "END"    : '\033[0m',
}


#class PublicationDB(SQLiteDB):
class PublicationDB(MongoDBConnector):
    """Class wrapping the SQLite database containing the publications.

    Parameters
    ----------
    filename : str
        Path to the SQLite database file.
    """
    def __init__(self, config=None):
        self.config = config
        #super().__init__(filename)
        super().__init__(self.config, 'kpub')

    def add(self, article, mission, snippits, instruments, archive, affiliation, reason, hasAcknowledgement):
        """Adds a single article object to the database.

        Parameters:
            article (json): Article json object returned from ADS API
            mission (str)
            snippits (JSON) : Snippits found in the article
            instruments (str): Pipe-delimited list of instruments
            archive (str): 0 or 1 indicating if archiving reference was found
            affiliation (str): 'keck', 'unrelated', or 'unknown'
            hasAcknowledgement (bool): True if Keck acknowledgement was found in the article snippets. 
        """
        log.debug('Ingesting {}'.format(article['bibcode']))

        # Store the extra metadata in the json string
        year, month = article['pubdate'][0:7].split('-')
        article['mission'] = mission
        article['instruments'] = instruments
        article['archive'] = archive
        self.add_row(article, month, year, mission, snippits, instruments, archive, affiliation, reason, hasAcknowledgement)


    def get_affiliation(self, snippits, mission):
        # Does snippits contain instrument strings and the mission is 'keck'? 
        # If so, then this is a Keck publication.
        
        acknowledgement = self.config.get('acknowledgement', [])
        keys = snippits.keys()
        affiliation = 'unknown' # default
        hasAcknowledgement = False
        reason = "Neither instr nor ack found."
        if any(x in keys for x in acknowledgement):
            reason = "Acknowledgement found in snippets."
            affiliation = 'keck' 
            hasAcknowledgement = True
        if len(snippits) > 0 and 'keck' in mission:
            reason = "Instrument names found in snippets."
            affiliation = 'keck' # pretty sure its keck
        if len(snippits) == 0 and not 'keck' in mission:
            reason = "No instrument names found in snippets."
            affiliation = 'unrelated' # pretty sure its unrelated
        return affiliation, hasAcknowledgement, reason

    def add_article(self, article, statusmsg="", interactive=False):
        """Adds an article via algorithm. the user can change the classification.

        Parameters:
            article (json): Article json object returned from ADS API
        """        
        # Do not show an article that is already in the database
        #if False:
        if self.article_exists(article):
            log.info("{} is already in the database "
                     "-- skipping.".format(article['bibcode']))
            #return 0


        # Print paper information to stdout
        #print(chr(27) + "[2J")  # Clear screen
        log.info(statusmsg)

        # Prompt the user to classify the paper by mission
        #NOTE: 'unrelated' is how things are permenantly marked to skip in DB.
        valmap = {'0': 'unrelated'}
        missions = self.config.get('missions', [])
        valmap = add_prompt_valmaps(valmap, missions)

        # snippits are instrument and 'keck' matches
        if interactive:
            log.info("\n([p] PDF view  [m] More context)")
            mission = prompt_grouping(valmap, 'Mission')
        else:
            mission = 'keck'

        if mission.lower() == 'm' or mission.lower() == 'keck':
            snippits = self.find_all_snippets(article['bibcode'])
        elif mission.lower() == 'p':
            self.open_pdf(article['bibcode'])

        instruments = ''
        if len(snippits) == 0:
            log.info("No snippets found.  Marking as unrelated.")
            mission = 'unrelated'
        instruments = "|".join([ x for x in snippits.keys() if x not in missions])

        # Get archive ack
        archive = self.get_archive_acknowledgement(snippits)

        # used for automation. Checks if this is a Keck publication.
        affiliation, hasAcknowledgement, reason = self.get_affiliation(snippits, mission)

        #add it
        self.add(article, mission=mission, 
                 snippits=snippits, instruments=instruments, 
                 archive=archive, affiliation=affiliation, 
                 reason=reason, hasAcknowledgement=hasAcknowledgement)
        return 1

    def find_all_snippets(self, bibcode):

        colors = self.config.get('colors')
        missions = self.config.get('missions', [])
        instruments = self.config.get('instruments', [])
        blacklist = self.config.get('blacklist', [])
        ads_api_key = self.config.get('ADS_API_KEY')
        acknowledgement = self.config.get('acknowledgement', [])
        archive = self.config.get('archive')

        #if not config for this, then return empty array
        words = []
        words += missions
        words += instruments
        words += acknowledgement
        words += archive
        if not words:
            return []

        #try two methods for finding matches

        try:
            counts = get_word_match_counts_by_pdf(bibcode, words, ads_api_key, blacklist)
        except Exception as err:
            log.warning(f"Could not parse PDF file. {err} Using alternate ADS query method...")
            counts = get_word_match_counts_by_query(bibcode, words, ads_api_key)

        #log.info snippets
        log.info("\nSNIPPETS FOUND:")
        for instr, count in counts.items():
            for snippet in count['snippets']:
                snippet = highlight_text(snippet, colors)
                log.info(f'"... {snippet}"')

        return counts


    def get_archive_acknowledgement(self, snippits):
        '''Search for instances of archive strings in full article.'''

        #if not config for this, then return empty array
        archive = self.config.get('archive')
        if not archive:
            return ''
        keys = snippits.keys()
        if any(x in keys for x in archive):
            log.info("Archive acknowledgement found in snippets.")
            return True 
        else: 
            return False 

    def set_affiliation(self, articles, affiliation, last_modifier='kpub', koa_affiliation=None):
        updated_articles = []
        for article in articles:
            # Get the bibcode
            article['date_modified'] = datetime.datetime.now()
            if koa_affiliation is not None:
                article['archive'] = koa_affiliation
            else:
                article['affiliation'] = affiliation
            article['last_modifier'] = last_modifier
            # Save the changes to the database
            updated_article = self.update_row_affiliation(article)
            if updated_article:
                updated_articles.append(updated_article)

        return updated_articles 


    def add_by_bibcode(self, bibcode, interactive=False):
        #TODO: NOTE: Without querying for 'keck' in full text, highlights will not be returned.
        bibcode = bibcode.replace('&', '%26')
        q = f"identifier:{bibcode}"
        data = self.query_ads(q)
        articles = data['response']['docs'] 

        if not articles:
            log.error(f"No ADS record found for bibcode {bibcode}")
        for article in articles:
            # Print useful warnings
            if bibcode != article['bibcode']:
                log.warning("Requested {} but ADS API returned {}".format(bibcode, article['bibcode']))
            if interactive and ('NONARTICLE' in article['property']):
                # Note: data products are sometimes tagged as NONARTICLE
                log.warning("{} is not an article.".format(article['bibcode']))
            else:
                self.add_article(article, interactive=interactive)

    def to_markdown(self, title="Publications",
                    group_by_month=False, save_as=None, **kwargs):
        """Returns the publication list in markdown format.
        """
        if group_by_month:
            group_idx = 'month' 
        else:
            group_idx = 'year'  # by year

        articles = collections.OrderedDict({})
        for row in self.query(**kwargs):
            group = str(row[group_idx])
            if group.endswith("-00"):
                group = group[:-3] + "-01"
            if group not in articles:
                articles[group] = []
            # The markdown template depends on "property" being iterable
            if row["property"] is None:
                row["property"] = []
            articles[group].append(row)

        templatedir = os.path.join(PACKAGEDIR, 'templates')
        env = jinja2.Environment(loader=jinja2.FileSystemLoader(templatedir))
        template = env.get_template('template.md')
        markdown = template.render(title=title, save_as=save_as,
                                   articles=articles)
        if sys.version_info >= (3, 0):
            return markdown  # Python 3
        else:
            return markdown.encode("utf-8")  # Python 2

    def save_markdown(self, output_fn, **kwargs):
        """Saves the database to a text file in markdown format.

        Parameters
        ----------
        output_fn : str
            Path of the file to write.
        """
        markdown = self.to_markdown(save_as=output_fn.replace("md", "html"),
                                    **kwargs)
        log.info('Writing {}'.format(output_fn))
        f = open(output_fn, 'w')
        f.write(markdown)
        f.close()

    def get_plot_data(self, plotname, **kwargs):
        """Returns the data for a given plot.

        Parameters
        ----------
        plotname : str
            Name of the plot to get data for.
        """
        year_begin = kwargs.get('year_begin', 2009)
        if year_begin is None: # Sometimes a None is passed
            year_begin = 2009

        if plotname == 'plot_by_year':
            extrapolate = kwargs.get('extrapolate', True)
            if extrapolate is None:
                extrapolate = True
            plotdata = plot.get_plot_by_year_data(self, year_begin=year_begin, extrapolate=extrapolate)
        elif plotname == 'plot_author_count':
            plotdata = plot.get_plot_author_count_data(self, year_begin=year_begin)
        elif plotname == 'plot_by_instrument':
            allInstruments = '|'.join(self.config.get('instruments', []))
            instruments = kwargs.get('instruments', allInstruments)
            if instruments is None:
                instruments = allInstruments
            instruments = instruments.split('|')
            plotdata, _ = plot.get_plot_instruments_data(self, year_begin=year_begin, instruments=instruments)
        else:
            raise ValueError(f"Unknown plot name: {plotname}") 

        print(f"Plot data for {plotname}: {plotdata}")
        return plotdata 

    def get_plot(self):
        """Saves beautiful plot of the database."""
        missions = self.config.get('missions', [])
        plots_cfg = self.config.get('plots', [])
        for ext in ['pdf', 'png']:
            plot.plot_by_year(self, f"{PLOTDIR}/kpub-publication-rate.{ext}", 
                              year_begin=plots_cfg['year_begin'])
            plot.plot_by_year(self, f"{PLOTDIR}/kpub-publication-rate-no-extrapolation.{ext}", 
                              year_begin=plots_cfg['year_begin'], extrapolate=False)
            plot.plot_author_count(self, f"{PLOTDIR}/kpub-author-count.{ext}", year_begin=plots_cfg['year_begin'])

        #bokeh plots
        if plots_cfg['instruments']:
            plot.plot_instruments(self, f"{PLOTDIR}/kpub-publications-by-instrument", 
                                  year_begin=plots_cfg['year_begin'],
                                  instruments=plots_cfg['instruments'])
        if self.config['aff_defs']:
            plot.plot_affiliations(self, f"{PLOTDIR}/kpub-affiliations", 
                                  year_begin=plots_cfg['year_begin'])


    def get_metrics(self, year=None):
        """Returns a dictionary of overall publication statistics.

        The metrics include:
        * # of publications since XX.
        * # of unique author surnames.
        * # of citations.
        * # of peer-reviewed pubs.
        """

        missions = self.config.get('missions', [])

        #init stats
        metrics = {}
        metrics['publication_count'] = 0
        metrics['refereed_count'] = 0
        metrics['citation_count'] = 0
        metrics['phd_count'] = 0
        for mission in [*missions, 'unknown', 'unrelated']:
            metrics[f'{mission}_count'] = 0
            metrics[f'{mission}_refereed_count'] = 0
            metrics[f'{mission}_citation_count'] = 0
            metrics[f'{mission}_phd_count'] = 0


        authors, first_authors = {}, {}
        authors['all'] = []
        first_authors['all'] = []
        for mission in missions:
            authors[mission] = []
            first_authors[mission] = []

        articles = self.query(year=year)
        for article in articles:

            #general count
            metrics["publication_count"] += 1
            metrics[f"{article['mission']}_count"] += 1

            #phd counts
            if "PhDT" in article['bibcode']:
                metrics["phd_count"] += 1
                metrics[f"{article['mission']}_phd_count"] += 1

            #author counts
            authors['all'].extend(article['author_norm'])
            first_authors['all'].append(article['first_author_norm'])
            authors[mission].extend(article['author_norm'])
            first_authors[mission].append(article['first_author_norm'])

            #refereed counts
            try:
                if "REFEREED" in article['property']:
                    metrics["refereed_count"] += 1
                    metrics[f"{article['mission']}_refereed_count"] += 1
            except TypeError:  # proprety is None
                pass

            #citation counts
            try:
                metrics["citation_count"] += article['citation_count']
                metrics[f"{article['mission']}_citation_count"] += article['citation_count']
            except (KeyError, TypeError):
                log.warning("{}: no citation_count".format(article["bibcode"]))

        metrics["author_count"] = np.unique(authors['all']).size
        metrics["first_author_count"] = np.unique(first_authors['all']).size
        for mission in missions:
            metrics[f"{mission}_author_count"] = np.unique(authors[mission]).size
            metrics[f"{mission}_first_author_count"] = np.unique(first_authors[mission]).size

        # Also compute fractions
        pubcount = metrics["publication_count"]
        for mission in missions:
            metrics[mission+"_fraction"] = metrics[mission+"_count"] / pubcount if pubcount > 0 else 0
    
        return metrics

    def get_all(self, mission=None):
        """Returns a list of dictionaries, one entry per publication."""
        return self.query(mission=mission)

    def get_most_cited(self, mission=None, top=10):
        """Returns the most-cited publications."""
        bibcodes, citations = [], []
        articles = self.query(mission=mission)
        for article in articles:
            bibcodes.append(article['bibcode'])
            if article["citation_count"] is None:
                citations.append(0)
            else:
                citations.append(article["citation_count"])
        idx_top = np.argsort(citations)[::-1][0:top]
        return [articles[idx] for idx in idx_top]

    def get_most_read(self, mission=None, top=10):
        """Returns the most-cited publications."""
        bibcodes, citations = [], []
        articles = self.query(mission=mission)
        for article in articles:
            bibcodes.append(article['bibcode'])
            citations.append(article["read_count"])
        idx_top = np.argsort(citations)[::-1][0:top]
        return [articles[idx] for idx in idx_top]

    def get_most_active_first_authors(self, min_papers=10):
        """Returns names and paper counts of the most active first authors."""
        articles = self.query()
        authors = {}
        for article in articles:
            first_author = article["first_author_norm"]
            try:
                authors[first_author] += 1
            except KeyError:
                authors[first_author] = 1
        names = np.array(list(authors.keys()))
        paper_count = np.array(list(authors.values()))
        idx_top = np.argsort(paper_count)[::-1]
        mask = paper_count[idx_top] >= min_papers
        return zip(names[idx_top], paper_count[idx_top[mask]])

    def get_all_authors(self, top=20):
        articles = self.query()
        authors = {}
        for article in articles:
            for auth in article["author_norm"]:
                try:
                    authors[auth] += 1
                except KeyError:
                    authors[auth] = 1
        names = np.array(list(authors.keys()))
        paper_count = np.array(list(authors.values()))
        idx_top = np.argsort(paper_count)[::-1][:top]
        return names[idx_top], paper_count[idx_top]

    def get_affiliation_counts(self, year_begin, year_end, mission):

        #init data
        counts = {}
        aff_defs = self.config['aff_defs']
        for affdef in aff_defs:
            counts['first author '+affdef['type']] = {}
            counts['top3 authors '+affdef['type']] = {}
            for year in range(year_begin, year_end+1):
                counts['first author '+affdef['type']][year] = 0
                counts['top3 authors '+affdef['type']][year] = 0

        articles = self.get_articles_by_mission_years(mission, year_begin, year_end)

        #for each article, get affiliations for first 3 authors for each article
        for article in articles:
            year = int(article['year'])
            num_affs = len(article['aff'])
            affs = []
            for i in range(0,3):
                if num_affs > i:
                    afftype = self.get_aff_type(article['aff'][i], aff_defs)
                    if not afftype: continue
                    affs.append(afftype)
                    if i == 0:
                        counts['first author '+afftype][year] += 1
            if len(affs) == 3 and len(set(affs)) == 1:
                counts['top3 authors '+afftype][year] += 1

        return counts

    def get_aff_type(self, affstr, aff_defs):
        '''
        Search for institution strings in affiliation string.  Affiliation string
        can have multiple semicolon-delimited entries.  'affmap' is an ordered 
        array of preferred affiliation types.  Each type has an array of strings to
        search for.
        '''
        #Sometimes the value is blank or "-"
        if len(affstr.strip()) <= 2:
            return None

        default = ''
        affs = affstr.split(";")
        for affdef in aff_defs:
            afftype = affdef['type']
            if not affdef['strings']: 
                default = afftype
                continue
            for string in affdef['strings']:
                for aff in affs:
                    if string.isupper():
                        if re.search(string, aff):
                            return afftype
                    else:
                        if re.search(string, aff, re.IGNORECASE):
                            return afftype                  
        return default

    def get_annual_publication_count(self, year_begin=2009, year_end=datetime.datetime.now().year,
                                     instrument=None):
        """Returns a dict containing the number of publications per year per mission.

        Parameters
        ----------
        year_begin : int
            Year to start counting. (default: 2009)

        year_end : int
            Year to end counting. (default: current year)
        """
        yeardict = self.get_articles_by_years_instrument(year_begin, year_end, instrument)
        return yeardict 

    def get_annual_publication_count_cumulative(self, year_begin=2009, year_end=datetime.datetime.now().year):
        """Returns a dict containing the cumulative number of publications per year.

        Parameters
        ----------
        year_begin : int
            Year to start counting. (default: 2009)

        year_end : int
            Year to end counting. (default: current year)
        """
        # Initialize a dictionary to contain the data to plot
        result = {}
        for year in range(year_begin, year_end + 1):
            cum = self.get_count_cumulative(str(year))
            result[year] = cum 
        # Also combine counts
        return result

    def update(self, month=None):
        """
        Query ADS for new publications.
        Parameters:
            month (str): Used for ADS pubdate param. Format "YYYY-MM" or "YYYY".
        """
        # # git pull reminder
        # log.info(HIGHLIGHTS['YELLOW'] +
        #       "Reminder: did you `git pull` kpub before running "
        #       "this command? [y/n] " +
        #       HIGHLIGHTS['END'],
        #       end='')
        # if input() == 'n':
        #     return

        #Assume current month if not supplied.
        #NOTE: We use the term "month" but user can supply just the year to do a whole year.
        if month is None:
            month = datetime.datetime.now().strftime("%Y-%m")

        #query 1
        queries = self.config.get('ads_queries')
        numArticlesAdded = 0
        for query in queries:
            log.info(f"\nQuerying {query['name']} (date={month})")
            data = self.query_ads(query['query'], month)
            articles = data.get('response', {}).get('docs', [])

            #loop and add
            for idx, article in enumerate(articles):

                # Ignore articles without abstract
                if not article.get('abstract'):
                    continue

                # Ignore proposals, cospar abstracts and tmp articles
                bibcode = article['bibcode']
                if ".prop." in bibcode or "cosp.." in bibcode or ".tmp" in bibcode:
                    continue

                # Propose to the user
                statusmsg = ("\n\n\n\n\n\n********** "
                    f"Showing article {idx+1} out of {len(articles)} ({query['name']} query)"
                    " **********\n")
                numArticlesAdded += self.add_article(article, statusmsg=statusmsg, interactive=False)

        #all done
        log.info(f'\nFinished reviewing all articles for {month}. added {numArticlesAdded} new articles.')
        return True, numArticlesAdded 

    def open_pdf(self, bibcode):
        '''Open PDF file in local browser.  Download if necessary.'''
        key = self.config.get('ADS_API_KEY')
        outfile = get_pdf_file(bibcode, key)
        if os.path.isfile(outfile):
            log.info(f"Opening {outfile}...")
            webbrowser.open('file://' + os.path.realpath(outfile))
            #webbrowser.get('firefox').open_new_tab('file://' + os.path.realpath(outfile))

    def query_ads(self, query, pubdate=None):
        '''
        Query ADS API.  Add in standard params needed for data store and text highlights.

        Parameters:
            query (str): An ADS compliant query string (exactly what is entered in web search GUI.)
            date (str): Optional ADS pubdate param. YYYY-MM or YYYY. Ex: "2019-03", "2020"
        '''

        query = query.replace(' ', '+')
        query = query.replace('"', '%22')
        if pubdate: query += f"+pubdate:{pubdate}"

        fl = ','.join(FIELDS)
        url = (f'{ADS_API}'
            f'q={query}'
            f"&fl={fl}"
            "&sort=date+asc"
            "&hl=true"
            "&hl.fl=ack,body,title,abstract"
            "&hl.snippets=4"
            "&hl.fragsize=100"
            "&hl.maxAnalyzedChars=500000"
            "&rows=9999999"
        )
        ads_api_key = self.config.get('ADS_API_KEY')
        data = request_ads_api(url, ads_api_key)
        return data 


##################
# Helper functions
##################

def request_ads_api(url, ads_api_key, returnResp=False):
    """Queries the ADS API with the given query string and returns the response data.

    Parameters
    ----------
    query : str
        The ADS query string to use for the API request.
    ads_api_key : str
        The API key for accessing the ADS API.

    Returns
    -------
    dict
        The JSON response from the ADS API.
    """
    headers = {'Authorization': f'Bearer {ads_api_key}'}
    resp = requests.get(url, headers=headers)
    rateLimitRem = resp.headers.get('X-RateLimit-Remaining', 100)
    if int(rateLimitRem) < 10:
        rateLimitReset = datetime.datetime.fromtimestamp(int(resp.headers.get('X-RateLimit-Reset')))
        log.warning(f"Rate limit remaining: {rateLimitRem}. Reset at {rateLimitReset} UTC.")
    resp.raise_for_status()
    if returnResp:
        return resp
    return resp.json()

def highlight_text(text, colors):

    for word, color in colors.items():
        pattern = re.compile(word, re.IGNORECASE)
        text = pattern.sub(HIGHLIGHTS[color] + word + HIGHLIGHTS['END'], str(text))
    return text
     
def display_abstract(article_dict, colors, highlights=None):
    """Prints the title and abstract of an article to the terminal,
    given a dictionary of the article metadata.

    Parameters
    ----------
    article : `dict` containing standard ADS metadata keys
    colors  : `dict` mapping keywords to colors
    highlights: `dict` containing 'ack' and 'body' lists of relevent text snippets
    """
    title = article_dict['title'][0]
    try:
        abstract = article_dict['abstract']
    except KeyError:
        abstract = ""

    title = highlight_text(title, colors)
    abstract = highlight_text(abstract, colors)

    ack_hl = 'NONE'
    if highlights and 'ack' in highlights:
        ack_hl = ''
        for ack in highlights['ack']:
            ack = ack.replace('<em>', '').replace('</em>', '')
            ack_hl += "\n\t" + '"...' + highlight_text(ack, colors) + '"'

    body_hl = 'NONE'
    if highlights and 'body' in highlights:
        body_hl = ''
        for body in highlights['body']:
            body = body.replace('<em>', '').replace('</em>', '')
            body_hl += "\n\t" + '"...' + highlight_text(body, colors) + '"'

    log.info(title)
    log.info('-'*len(title))
    log.info(abstract)
    log.info('')
    log.info(f"Acknowledgement highlights: {ack_hl}")
    log.info(f"Body highlights: {body_hl}")
    log.info('')
    log.info('Authors: ' + ', '.join(article_dict.get('author', '')))
    log.info('Date: ' + article_dict['pubdate'])
    log.info('Status: ' + str(article_dict['property']))
    log.info('URL: http://adsabs.harvard.edu/abs/' + article_dict['bibcode'])
    log.info('')


def get_word_match_counts_by_query(bibcode, words, ads_api_key):

    bibcode = bibcode.replace('&', '%26')

    counts = {}
    for word in words:
        word = word.replace(' ', '+')
        url = (f'{ADS_API}' 
            f'q=bibcode:%22{bibcode}%22+full:%22{word}%22'
            "&fl=id,bibcode"
            "&sort=date+asc"
            "&hl=true"
            "&hl.fl=ack,body,title,abstract"
            "&hl.snippets=4"
            "&hl.fragsize=100"
            "&hl.maxAnalyzedChars=500000"
        )
        data = request_ads_api(url, ads_api_key)
        counts[word] = {'count': 0, 'snippets': []}
        for doc in data.get('response', {}).get('docs',[]):
            id = doc['id']
            highlights = data['highlighting'][id]
            for _, snippets in highlights.items():
                counts[word]['count'] += len(snippets)
                counts[word]['snippets'] = snippets

    #only return counts > 0
    counts = {key:val for key, val in counts.items() if val['count'] != 0}
    return counts
 

def get_word_match_counts_by_pdf(bibcode, words, ads_api_key, blacklist=[]):

    #get pdf file and text
    outfile = get_pdf_file(bibcode, ads_api_key)
    #text = get_pdf_text(outfile).lower()
    text = get_pdf_text(outfile)
    text = text.replace("\n",' ')
    text = text.replace("\r",' ')
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\xff]', ' ', text)

    #count up matches
    counts = {}
    for word in words:
        counts[word] = {'count': 0, 'snippets': []}
        for ch in (' ', '/', '\(', '-', ':'):
            #find = f"{ch}{word}".lower()
            find = f"{ch}{word}"
            for match in re.finditer(find, text):
                #skip if text in blacklist
                snip = text[match.start()-5:match.end()+5]
                if any(bl in snip for bl in blacklist):
                    continue
                snippet = text[match.start()-80:match.end()+80]
                counts[word]['count'] += 1
                counts[word]['snippets'].append(snippet)

    #only return counts > 0
    counts = {key:val for key, val in counts.items() if val['count'] != 0}
    return counts
  

def get_pdf_file(bibcode, ads_api_key):

    outfile = f'/tmp/{bibcode}.pdf'
    if os.path.isfile(outfile):
        return outfile

    log.info('\nRetrieving PDF (May take up to a minute)...')
    url = f'https://ui.adsabs.harvard.edu/link_gateway/{bibcode}/EPRINT_PDF'
    resp = request_ads_api(url, ads_api_key, returnResp=True)
    if resp.status_code != 200 or len(resp.content) < 1000:
        print("Could not download PDF file.")
        return False
    with open(outfile, 'wb') as f:
         f.write(resp.content)
    return outfile


def get_pdf_text(outfile):
    assert textract, "No textract module found."
    methods = ['pdftotext', 'pdfminer']
    text = ''
    for method in methods:
        try:
            text = textract.process(outfile, method=method)
            text = text.decode("utf-8")
            if text: return text
        except Exception as e:
            log.info(f"textract: {method} method failed.  Trying another method...")
    if not text:
        raise Exception("Could not extract PDF text")


def input_with_prefill(prompt, text):
    def hook():
        readline.insert_text(text)
        readline.redisplay()
    readline.set_pre_input_hook(hook)
    result = input(prompt)
    readline.set_pre_input_hook()
    return result


def add_prompt_valmaps(valmap, vals):

    for idx, val in enumerate(vals):
        k = str(idx+1)
        valmap[k] = val
    return valmap


def prompt_grouping(valmap, type):

    prompt = f"=> Select {type}: "
    for key, val in valmap.items():
        prompt += f" [{key}] {val.capitalize()} "
    prompt += " or [] skip? "

    log.info(prompt)
    val = input()
    return val




#########################
# Command-line interfaces
#########################

def kpub_stats():
    """Save the publication stats in Markdown format."""

    config = yaml.load(open(f'{PACKAGEDIR}/config/config.live.yaml'), Loader=yaml.FullLoader)
    title = config.get('prepend', '').capitalize()

    pubdb = PublicationDB(config)

    for bymonth in [True, False]:
        if bymonth:
            suffix = "-by-month"
            title_suffix = " by month"
        else:
            suffix = ""
            title_suffix = ""

        output_fn = f"{MDDIR}/kpub-{config['prepend']}-publications{suffix}.md"
        pubdb.save_markdown(output_fn,
                         group_by_month=bymonth,
                         title=f"{title} publications{title_suffix}")

    # Finally, produce an overview page
    templatedir = os.path.join(PACKAGEDIR, 'templates')
    env = jinja2.Environment(loader=jinja2.FileSystemLoader(templatedir))
    template = env.get_template('template-overview.md')
    markdown = template.render(institution=title,
                               metrics=pubdb.get_metrics(),
                               most_cited=pubdb.get_most_cited(top=20),
                               most_active_first_authors=pubdb.get_most_active_first_authors(),
                               now=datetime.datetime.now())
    # most_read=pubdb.get_most_read(20),
    filename = f'{MDDIR}/publications-overview.md'
    log.info('Writing {}'.format(filename))
    f = open(filename, 'w')
    if sys.version_info >= (3, 0):
        f.write(markdown)  # Python 3
    else:
        f.write(markdown.encode("utf-8"))  # Legacy Python
    f.close()

def kpub_plot_data(plotname, instruments=None, year_begin=None, extrapolate=False):
    """Creates beautiful data for plotting."""
    config = yaml.load(open(f'{PACKAGEDIR}/config/config.live.yaml'), Loader=yaml.FullLoader)
    pubdb = PublicationDB(config)
    plotData = pubdb.get_plot_data(plotname=plotname, instruments=instruments, extrapolate=extrapolate, year_begin=year_begin)
    return plotData

def kpub_plot():
    """Creates beautiful plots of the database."""
    config = yaml.load(open(f'{PACKAGEDIR}/config/config.live.yaml'), Loader=yaml.FullLoader)
    pubdb = PublicationDB(config)
    pubdb.get_plot()

def kpub_update(month):
    """Interactively query ADS for new publications."""

    config = yaml.load(open(f'{PACKAGEDIR}/config/config.live.yaml'), Loader=yaml.FullLoader)

    db = PublicationDB(config)
    success, numArticlesAdded = db.update(month=month)
    return success, numArticlesAdded 

def kpub_add(bibcodes, interactive=False):
    """Add a publication with a known ADS bibcode."""

    config = yaml.load(open(f'{PACKAGEDIR}/config/config.live.yaml'), Loader=yaml.FullLoader)

    db = PublicationDB(config)
    for bibcode in bibcodes:
        db.add_by_bibcode(bibcode, interactive=interactive)


def kpub_delete(bibcodes):
    """Deletes a publication using its ADS bibcode."""

    config = yaml.load(open(f'{PACKAGEDIR}/config/config.live.yaml'), Loader=yaml.FullLoader)

    db = PublicationDB(config)
    for bibcode in bibcodes:
        db.delete_by_bibcode(bibcode)


def kpub_import(jsonfile):
    """Import publications from a json file.

    The json file must contain entries of the form "bibcode,mission".
    The actual metadata of each publication will be grabbed using the ADS API,
    hence this routine may take 10-20 minutes to complete.
    """
    config = yaml.load(open(f'{PACKAGEDIR}/config/config.live.yaml'), Loader=yaml.FullLoader)

    db = PublicationDB(config)
    with open(jsonfile, 'r') as f:
        rows = json.load()
    for row in rows:
        try:
            bibcode = row['bibcode'] 
            mission = row['mission'] 
            instrs  = row['instruments'] 
            archive = row['archive'] 
            db.add_by_bibcode(bibcode, mission=mission,
                instruments=instrs, archive=archive, interactive=False)
        except Exception as err:
            log.warning("attempt #{} for {}: error '{}'".format(row, err))

    #all done
    log.info(f'\nFinished importing.')
    log.info(HIGHLIGHTS['YELLOW'] +
          "\nREMINDER: Do a `kpub push` to update the data files in github!" +
          HIGHLIGHTS['END'])

def kpub_set_affiliation(articles, affiliation, last_modifier, koa_affiliation=None):

    config = yaml.load(open(f'{PACKAGEDIR}/config/config.live.yaml'), Loader=yaml.FullLoader)
    db = PublicationDB(config)
    articles = db.set_affiliation(articles, affiliation, last_modifier, koa_affiliation=koa_affiliation)
    log.info('Set affiliation for {} articles to {}'.format(len(articles), affiliation))
    return articles
    

def kpub_export(monthyear, begin_year=None, filename=None, affiliation=None, csv=None):
    """Export the database as JSON format."""
    config = yaml.load(open(f'{PACKAGEDIR}/config/config.live.yaml'), Loader=yaml.FullLoader)
    db = PublicationDB(config)
    year, month = monthyear.split('-') if '-' in monthyear else (monthyear, None)
    year, month = int(year), int(month) if month else None
    begin_year = int(begin_year) if begin_year else None
    articles = db.get_articles(begin_year=begin_year, end_year=year, month=month, affiliation=affiliation)
    if not articles:
        log.info('No rows found.')
        return []
    # Convert to a list of dictionaries
    if not filename:
        log.info('No filename specified.  Returning articles as list of dicts.')
        return articles

    if csv:
        df = pd.DataFrame(articles)
        df.to_csv(filename, index=False)
    else:
        with open(filename, 'w') as f:
            json.dump(articles, f, indent=4, default=serialize_datetime)
    log.info(f'Wrote {len(articles)} articles to {filename}')
    
def serialize_datetime(obj):
    if isinstance(obj, datetime.datetime):
        return obj.isoformat()
    raise TypeError("Type not serializable")

def kpub_spreadsheet(filename):
    """Export the publication database to an Excel spreadsheet."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        log.error('openpyxl needs to be installed for this feature.')
        sys.exit(1)

    config = yaml.load(open(f'{PACKAGEDIR}/config/config.live.yaml'), Loader=yaml.FullLoader)

    db = PublicationDB(filename, config)
    spreadsheet = []
    rows = db.select_for_spreadsheet()
    for row in rows:
        try:
            if 'REFEREED' in row['property']:
                refereed = 'REFEREED'
            elif 'NOT REFEREED' in row['property']:
                refereed = 'NOT REFEREED'
            else:
                refereed = ''
        except TypeError:  # .property is None
            refereed = ''
        # Compute citations per year
        try:
            dateobj = datetime.datetime.strptime(row['date'], '%Y-%m-00')
        except ValueError:
            dateobj = datetime.datetime.strptime(row['date'], '%Y-00-00')
        publication_age = datetime.datetime.now() - dateobj
        try:
            citations_per_year = row['citation_count'] / (publication_age.days / 365)
        except (TypeError, ZeroDivisionError):
            citations_per_year = 0

        myrow = collections.OrderedDict([
                    ('bibcode', row['bibcode']),
                    ('year', row['year']),
                    ('date', row['date']),
                    ('mission', row['mission']),
                    ('date_modified', row['date_modified']),
                    ('last_modifier', row['last_modifier']),
                    ('refereed', refereed),
                    ('citation_count', row['citation_count']),
                    ('citations_per_year', round(citations_per_year, 2)),
                    ('read_count', str(row['read_count'])),
                    ('first_author_norm', str(row['first_author_norm'])),
                    ('title', row['title'][0]),
                    ('keyword_norm', str(row.get('keyword_norm'))),
                    ('keyword', str(row.get('keyword'))),
                    ('abstract', row['abstract']),
                    ('co_author_norm', str(row['author_norm'])),
                    ('instruments', str(row['instruments'])),
                    ('archive', str(row['archive'])),
                    ('affiliation', str(row['affiliation'])),
                    ('affiliations', str(row['aff']))
                    ])
        spreadsheet.append(myrow)

    output_fn = 'kpub-publications.xlsx'
    log.info('Writing to {}'.format(output_fn))
    wb = Workbook()
    ws = wb.active
    fieldnames = list(spreadsheet[0].keys())
    ws.append(fieldnames)
    for row in spreadsheet:
        values = [row[k] for k in fieldnames]
        ws.append(values)
    for cell in ws['A'] + ws[1]:
        cell.style = 'Pandas'
    wb.save(output_fn)

def make_parser():

    parser = argparse.ArgumentParser(
        description="kpub: a tool to manage the publication database.")
    subparsers = parser.add_subparsers(dest='command')
    subparsers.required = True
    update_parser = subparsers.add_parser('update', help='Update the database with new publications.')
    update_parser.add_argument('month', nargs='?', default=None,
                        help='Month to query, YYYY-MM or YYYY. e.g. "2015-06" or "2020"')

    plot_parser = subparsers.add_parser('plot', help='Plot the publication data.')

    plot_data_parser = subparsers.add_parser('plot_data', help='Get the data for a plot.')
    plot_data_parser.add_argument('plotname', nargs='?', default='plot_by_year', type=str,
                        help='Name of the plot to get data for.')
    plot_data_parser.add_argument('instruments', nargs='?', default=None,
                        help='Instruments to plot. Pipe separated list. e.g. "ESI|HIRES|NIRSPEC"')
    plot_data_parser.add_argument('year_begin', nargs='?', default=None,
                        help='year to begin the data collection. e.g. "2015"')
    plot_data_parser.add_argument('extrapolate', nargs='?', default=False,
                        help='Extrapolate the data to the current date.')

    add_parser = subparsers.add_parser('add', help='Add a publication to the database.')
    add_parser.add_argument('bibcode', nargs='+',
                        help='ADS bibcode that identifies the publication.')
    add_parser.add_argument('-interactive', action='store_true',
                        help='Interactive mode.  Prompt for each article.')

    delete_parser = subparsers.add_parser('delete', help='Delete a publication from the database.')
    delete_parser.add_argument('bibcode', nargs='+',
                        help='ADS bibcode that identifies the publication.')
    
    import_parser = subparsers.add_parser('import', help="Batch-import papers into the publication list "
                    "from a JSON file. The JSON file must have bibcode"
                    "For example: '2004ApJ...610.1199G,kepler,astrophysics'.")
    import_parser.add_argument('jsonfile',
                        help="Filename of the JSON file to ingest.")

    export_parser = subparsers.add_parser('export', help='Batch-export articles to a JSON file (or CSV if specified).')
    export_parser.add_argument('monthyear', type=str, nargs='?', default=datetime.datetime.now().strftime("%Y-%m"),
                        metavar='YYYY-MM',
                        help="Month Year to export. YYYY-MM or YYYY. e.g. '2015-06' or '2020'")
    export_parser.add_argument('begin_year', type=int, nargs='?',
                        help="Begining year to export. (if range is desired)")
    export_parser.add_argument('affiliation', type=str, nargs='?',
                        help="Affiliation to export.")
    export_parser.add_argument('-filename', type=str, nargs='?',
                        help="Filename to export to.")
    export_parser.add_argument('-csv', action='store_true',
                        help="Export as CSV instead of JSON.")

    stats_parser = subparsers.add_parser('stats', help='Get the publication statistics.')

    spreadsheet_parser=subparsers.add_parser('spreadsheet', help='Export the database to a spreadsheet.')
    spreadsheet_parser.add_argument('filename', type=str, help='Filename of the spreadsheet to export to.')

    return parser

if __name__ == "__main__":

    cmd = sys.argv[1]
    parser = make_parser()
    margs = parser.parse_args(sys.argv[1:])
    if cmd == 'update':      kpub_update(margs.month)
    elif cmd == 'add':       kpub_add(margs.bibcode, margs.interactive)
    elif cmd == 'plot':      kpub_plot()
    elif cmd == 'plot_data': kpub_plot_data(margs.plotname, margs.instruments, margs.year_begin, margs.extrapolate)
    elif cmd == 'delete':    kpub_delete(margs.bibcode)
    elif cmd == 'import':    kpub_import(margs.jsonfile)
    elif cmd == 'export':    kpub_export(margs.monthyear, margs.begin_year, margs.filename, margs.affiliation, margs.csv)
    elif cmd == 'stats':     kpub_stats()
    elif cmd == 'spreadsheet': kpub_spreadsheet(margs.filename)
    else: log.error("Unknown kpub command")
